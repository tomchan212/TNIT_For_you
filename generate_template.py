#!/usr/bin/env python3
"""Generate HKCYSTINTJustForYou Google Sheet / Excel template."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "HKCYSTINTJustForYou_Sheet_Template.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="C4A574", end_color="C4A574", fill_type="solid")
NOTE_FONT = Font(italic=True, color="8B7355", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="E8DFD0"),
    right=Side(style="thin", color="E8DFD0"),
    top=Side(style="thin", color="E8DFD0"),
    bottom=Side(style="thin", color="E8DFD0"),
)

def build_participants():
    rows = []
    phones = {
        "1A": "98765432", "2A": "98765433", "3A": "98765434", "4A": "98765435",
        "5A": "98765436", "6A": "98765437", "1B": "98765438", "2B": "98765439",
        "3B": "98765440", "4B": "98765441", "5B": "98765442", "6B": "98765443",
        "1C": "98765444", "2C": "98765445", "3C": "98765446", "4C": "98765447",
        "5C": "98765448", "6C": "98765449", "1D": "98765450", "2D": "98765451",
        "3D": "98765452", "4D": "98765453", "5D": "98765454", "6D": "98765455",
        "1E": "98765456", "2E": "98765457", "3E": "98765458", "4E": "98765459",
        "5E": "98765460", "6E": "98765461", "1F": "98765462", "2F": "98765463",
        "3F": "98765464", "4F": "98765465", "5F": "98765466", "6F": "98765467",
        "1G": "98765468", "2G": "98765469", "3G": "98765470", "4G": "98765471",
        "5G": "98765472", "6G": "98765473", "1H": "98765474", "2H": "98765475",
        "3H": "98765476", "4H": "98765477", "5H": "98765478", "6H": "98765479",
        "WILL": "11111111", "PHOEBE": "22222222", "RIGHT": "33333333",
        "RENAE": "44444444", "CHUNJAI": "55555555", "IRENE": "66666666",
        "KAKA": "77777777", "KAYI": "88888888", "WINGLAM": "99999999",
        "CHERYL": "00000000", "QQ": "12121212", "KELLY": "13131313", "BRO": "14141414",
    }
    for n in range(1, 7):
        for letter in "ABCDEFGH":
            pid = f"{n}{letter}"
            rows.append([pid, phones.get(pid, ""), f"GROUP_{n}"])
    for pid in ["WILL", "PHOEBE", "RIGHT", "RENAE", "CHUNJAI", "IRENE",
                "KAKA", "KAYI", "WINGLAM", "CHERYL", "QQ", "KELLY", "BRO"]:
        rows.append([pid, phones.get(pid, ""), "GROUP_STAFF"])
    return rows


def style_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 4)


def add_note(ws, row, text, merge_cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True)


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # ── Participants ──
    ws = wb.create_sheet("Participants")
    headers = ["participant_id", "phone_number", "group_id"]
    style_header(ws, headers)
    samples = build_participants()
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val).border = THIN_BORDER
    add_note(ws, 64, "說明：group_id 必須相同才能成為 Trophy 隊友。1A-1H 用 GROUP_1，2A-2H 用 GROUP_2… 工作人員用 GROUP_STAFF。勿將電話號碼填入 group_id！")

    # ── Messages ──
    ws = wb.create_sheet("Messages")
    headers = ["message_id", "sender_id", "receiver_id", "content", "created_at", "status", "deleted_at"]
    style_header(ws, headers)
    add_note(ws, 3, "說明：由系統自動寫入。status = active 或 deleted。收件箱不顯示 deleted 訊息。")

    # ── Open ──
    ws = wb.create_sheet("Open")
    ws["A1"] = "messaging_status"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A2"] = "OPEN"
    ws.column_dimensions["A"].width = 20
    add_note(ws, 4, "說明：A2 填入 OPEN（開放留言）或 CLOSE（關閉留言）。管理員可從控制台切換。")

    # ── Trophy ──
    ws = wb.create_sheet("Trophy")
    headers = ["Trophy_id", "Trophy_name"]
    style_header(ws, headers)
    trophies = [
        ["T01", "氣球人"],
        ["T02", "破冰人"],
        ["T03", "和事佬"],
        ["T04", "真好人"],
        ["T05", "組爸媽繼承人"],
        ["T06", "關鍵人物"],
        ["T07", "潛行者"],
        ["T08", "頭頭是道"],
    ]
    for r, row in enumerate(trophies, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val).border = THIN_BORDER
    add_note(ws, 10, "說明：定義所有可分配的 Trophy 稱號。Trophy_id 需唯一。")

    # ── Trophy_log ──
    ws = wb.create_sheet("Trophy_log")
    headers = ["Tmessage_id", "sender_id", "receiver_id", "Trophy_id"]
    style_header(ws, headers)
    add_note(ws, 3, "說明：參加者提交後的正式投票記錄。由系統寫入。")

    # ── Trophy_draft ──
    ws = wb.create_sheet("Trophy_draft")
    headers = ["Tmessage_id", "sender_id", "receiver_id", "Trophy_id"]
    style_header(ws, headers)
    add_note(ws, 3, "說明：參加者儲存的草稿。提交後會移至 Trophy_log 並清除草稿。")

    # ── Trophy_submissions ──
    ws = wb.create_sheet("Trophy_submissions")
    headers = ["participant_id", "submission_status", "submitted_at", "updated_at"]
    style_header(ws, headers)
    add_note(ws, 3, "說明：submission_status = draft 或 submitted。由系統寫入。")

    # ── Trophy_results ──
    ws = wb.create_sheet("Trophy_results")
    headers = ["participant_id", "Trophy_id", "award_source", "calculated_at"]
    style_header(ws, headers)
    add_note(ws, 3, "說明：管理員計算結果後寫入。award_source = round1（組內最高票）。")

    # ── Voting ──
    ws = wb.create_sheet("Voting")
    labels = ["voting_status", "allow_resubmit", "calculated_at", "published_at"]
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws["A2"] = "DRAFT"
    ws["B2"] = "FALSE"
    ws["C2"] = ""
    ws["D2"] = ""
    add_note(ws, 4, "說明：A2 = DRAFT | VOTING_OPEN | VOTING_CLOSED | CALCULATED | PUBLISHED。B2 = TRUE/FALSE（是否允許重填）。")

    # ── README ──
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    readme_rows = [
        ("HKCYS TINT — Just For You", ""),
        ("", ""),
        ("分頁名稱", "用途"),
        ("Participants", "參加者名單（登入驗證 + Trophy 分組）"),
        ("Messages", "匿名留言記錄"),
        ("Open", "全域留言開關（A2: OPEN / CLOSE）"),
        ("Trophy", "Trophy 稱號定義"),
        ("Trophy_log", "正式投票記錄"),
        ("Trophy_draft", "投票草稿"),
        ("Trophy_submissions", "各參加者提交狀態"),
        ("Trophy_results", "計算後的 Trophy 結果"),
        ("Voting", "投票生命週期設定"),
        ("", ""),
        ("管理員登入", "participant_id = admin，phone = 23082026"),
        ("", ""),
        ("部署步驟", ""),
        ("1", "上傳此檔案到 Google Drive → 以 Google 試算表開啟"),
        ("2", "確認各分頁名稱完全一致（區分大小寫）"),
        ("3", "在 Participants / Trophy 填入實際資料"),
        ("4", "Extensions → Apps Script → 貼上 Code.gs → 綁定試算表"),
        ("5", "Deploy → Web App（Execute as Me, Anyone）"),
        ("6", "將部署 URL 填入 app.js 的 API_URL"),
    ]
    title_font = Font(bold=True, size=14, color="5C4A38")
    for r, (a, b) in enumerate(readme_rows, 1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if r == 1:
            ws.cell(row=r, column=1).font = title_font
        if r == 3:
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=2).font = Font(bold=True)

    return wb


if __name__ == "__main__":
    wb = build_workbook()
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")
